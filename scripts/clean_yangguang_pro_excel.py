"""Clean Yangguang Pro survey and preserve model-reason relationships."""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

D = "\u250b"
SKIP = {"", "(\u8df3\u8fc7)", "\u8df3\u8fc7", "\uff08\u8df3\u8fc7\uff09"}
MODEL_RULES = [
    ("\u4e94\u83f1\u626c\u5149\u5ba2\u8f66\u7248", r"\u4e94\u83f1\u626c\u5149\u5ba2\u8f66"),
    ("\u4e94\u83f1\u626c\u5149\u5c01\u7a97\u8f66", r"\u4e94\u83f1\u626c\u5149\u5c01\u7a97"),
    ("\u8fdc\u7a0b\u661f\u4eabV6E\u5c01\u7a97\u8f66", r"\u661f\u4eabV6E"),
    ("\u957f\u5b89\u8de8\u8d8a\u661fV7EV\u5ba2\u8f66\u7248", r"\u8de8\u8d8a\u661fV7EV"),
    ("\u6c5f\u94c3E\u798f\u987a\u957f\u8f74\u4e2d\u9876\u5ba2\u8f66", r"E\u798f\u987a\u957f\u8f74\u4e2d\u9876\u5ba2\u8f66"),
    ("\u6c5f\u94c3E\u798f\u987a\u957f\u8f74\u4f4e\u9876\u5ba2\u8f66", r"E\u798f\u987a\u957f\u8f74\u4f4e\u9876\u5ba2\u8f66"),
    ("\u6c5f\u94c3E\u798f\u987a\u957f\u8f74\u4e2d\u9876\u5c01\u7a97\u8f66", r"E\u798f\u987a\u957f\u8f74\u4e2d\u9876\u5c01\u7a97"),
    ("\u6c5f\u94c3E\u798f\u987a\u957f\u8f74\u4f4e\u9876\u5c01\u7a97\u8f66", r"E\u798f\u987a\u957f\u8f74\u4f4e\u9876\u5c01\u7a97"),
    ("\u6c5f\u94c3\u798f\u987a\u5ba2\u8f66", r"\u6c5f\u94c3\u798f\u987a\u5ba2\u8f66"),
    ("\u6c5f\u94c3\u798f\u987a\u5c01\u7a97\u8f66", r"\u6c5f\u94c3\u798f\u987a\u5c01\u7a97"),
    ("\u8fdc\u7a0b\u8d85\u7ea7VAN", r"\u8fdc\u7a0b\u8d85\u7ea7VAN"),
    ("\u957f\u5b89V919", r"\u957f\u5b89V919"), ("\u5927\u901aV80\u5ba2\u8f66", r"\u5927\u901aV80\u5ba2\u8f66"),
    ("\u5927\u901aV80\u5c01\u7a97\u8f66", r"\u5927\u901aV80\u5c01\u7a97"), ("\u5927\u901aG10", r"\u5927\u901aG10"),
    ("\u591a\u62c9\u5927\u9762", r"\u591a\u62c9\u5927\u9762"), ("\u738b\u724cM7", r"\u738b\u724cM7"),
]
REASONS = [
    ("\u7a7a\u95f4/\u88c5\u8f7d", r"\u7a7a\u95f4|\u5bbd|\u7a84|\u5c0f|\u5927|\u88c5|\u62c9\u8d27|\u591a\u62c9|\u627f\u8f7d|\u8f7d\u91cd|\u5428|\u80fd\u88c5"),
    ("\u4ef7\u683c/\u8d39\u7528", r"\u4ef7\u683c|\u8f66\u4ef7|\u8d35|\u6027\u4ef7\u6bd4|\u6210\u672c|\u4fdd\u9669|\u91d1\u878d"),
    ("\u54c1\u724c/\u53e3\u7891/\u552e\u540e", r"\u54c1\u724c|\u724c\u5b50|\u53e3\u7891|\u4fe1\u4efb|\u653e\u5fc3|\u552e\u540e|\u95e8\u5e97"),
    ("\u52a8\u529b/\u5e95\u76d8/\u64cd\u63a7", r"\u52a8\u529b|\u626d\u77e9|\u540e\u9a71|\u524d\u9a71|\u5239\u8f66|\u5e95\u76d8|\u8f6e\u80ce"),
    ("\u7eed\u822a/\u7535\u6c60/\u8865\u80fd", r"\u7eed\u822a|\u91cc\u7a0b|\u7535\u6c60|\u7535\u673a|\u5145\u7535|2C|\u98ce\u51b7|\u8fbe\u6210\u7387"),
    ("\u8def\u6743/\u8f66\u578b\u5c5e\u6027", r"\u9650\u884c|\u8fdb\u57ce|\u5c01\u7a97|\u5ba2\u8f66|\u6cb9\u8f66|\u71c3\u6cb9"),
    ("\u4ea7\u54c1/\u914d\u7f6e/\u5916\u89c2", r"\u914d\u7f6e|\u5916\u89c2|\u597d\u770b|\u524d\u8138|\u8f66\u7a97|\u73bb\u7483|\u653e\u7535|\u592a\u9ad8"),
    ("\u9500\u552e\u670d\u52a1", r"\u9500\u552e|\u8ddf\u8e2a|\u4e13\u4e1a"),
]
CLEAN_HEADERS = [
    "\u6837\u672c\u5e8f\u53f7", "\u63d0\u4ea4\u7b54\u5377\u65f6\u95f4", "\u6240\u7528\u65f6\u95f4", "\u6765\u6e90", "\u6765\u6e90\u8be6\u60c5", "\u6765\u81eaIP",
    "\u624b\u673a\u53f7", "\u6240\u5728\u5730\u533a", "\u626c\u5149Pro\u8f66\u578b\u914d\u7f6e", "\u8f66\u8f86\u4e3b\u8981\u4f7f\u7528\u4eba", "\u8d2d\u8f66\u65b9\u5f0f", "\u4e86\u89e3\u6e20\u9053",
    "\u4e0a\u7f51\u4e60\u60ef", "\u5145\u7535\u573a\u666f", "\u8d27\u8fd0\u7528\u6237\u7c7b\u578b", "\u8d27\u8fd0\u884c\u4e1a\u5206\u7c7b", "\u5e38\u9a7b\u4f5c\u4e1a\u573a\u5730", "\u884c\u4e1a\u4fe1\u606f\u53ca\u8d2d\u8f66\u798f\u5229\u6e20\u9053",
    "\u56fa\u5b9a\u540c\u884c\u53f8\u673a\u8f66\u4e3b\u5708\u89c4\u6a21", "\u8f66\u4e3b\u5708\u670d\u52a1\u65b9\u5f0f", "\u65e5\u5747\u884c\u9a76\u91cc\u7a0b", "\u8d2d\u4e70\u5f62\u5f0f", "\u539f\u6709\u8f66\u578b", "\u5bf9\u6bd4\u8f66\u578b\u4e0e\u653e\u5f03\u539f\u56e0\uff08\u539f\u59cb\uff09",
    "\u65b0\u80fd\u6e90\u5546\u7528\u8f66\u5173\u6ce8\u70b9", "\u8d2d\u4e70\u65b0\u80fd\u6e90\u5546\u7528\u8f66\u539f\u56e0", "\u8d2d\u4e70\u626c\u5149Pro\u4e3b\u8981\u539f\u56e0", "\u8f66\u4e3b\u5408\u4f19\u4eba\u610f\u613f", "\u8f6c\u4ecb\u7ecd\u5956\u52b1\u504f\u597d", "\u7ebf\u4e0b\u6d3b\u52a8\u65b9\u4fbf\u65f6\u6bb5",
    "\u7ebf\u4e0b\u6d3b\u52a8\u504f\u597d", "\u4e0a\u95e8\u5173\u6000\u793c\u54c1\u504f\u597d", "\u4e2a\u4eba\u7167\u7247",
]

def txt(v: Any) -> str:
    return "" if v is None else str(v).strip()

def clean_value(v: Any) -> str:
    s = txt(v)
    return "" if s in SKIP else D.join(dict.fromkeys(x.strip() for x in s.split(D) if x.strip() not in SKIP))

def model_name(s: str) -> str:
    return next((name for name, pat in MODEL_RULES if re.search(pat, s, re.I)), s or "\u672a\u6ce8\u660e\u8f66\u578b")

def reason_type(s: str) -> str:
    if not s or s in {"\u65e0", "\u6ca1\u6709", "\u5fd8\u4e86", "\u597d"}:
        return "\u672a\u8bf4\u660e/\u65e0\u6709\u6548\u539f\u56e0"
    return next((name for name, pat in REASONS if re.search(pat, s, re.I)), "\u5176\u4ed6\u660e\u786e\u539f\u56e0")

def parse(v: Any) -> list[tuple[str, str, str, str]]:
    out = []
    for item in clean_value(v).split(D):
        m = re.fullmatch(r"(.*?)\u3016(.*?)\u3017", item)
        car, why = (m.group(1).strip(), m.group(2).strip()) if m else (item.strip(), "")
        quality = "\u5b8c\u6574" if why else "\u539f\u56e0\u672a\u586b\u5199"
        if car == "\u5176\u4ed6\u8f66\u578b":
            if re.fullmatch(r"\u5927\u62ffV1|\u9ec4\u91d1\u5361[\uff0c,]?", why):
                car, why, quality = why.rstrip("\uff0c,"), "", "\u5176\u4ed6\u8f66\u578b\u540d\u5df2\u63d0\u53d6\uff1b\u539f\u56e0\u672a\u586b\u5199"
            elif why.startswith("\u5409\u5229\u8fdc\u7a0bV8E"):
                car, why, quality = "\u5409\u5229\u8fdc\u7a0bV8E", why.removeprefix("\u5409\u5229\u8fdc\u7a0bV8E"), "\u5176\u4ed6\u8f66\u578b\u540d\u5df2\u63d0\u53d6"
            elif why.startswith("\u798f\u7530"):
                car, why, quality = "\u798f\u7530\uff08\u8f66\u578b\u672a\u6ce8\u660e\uff09", why.removeprefix("\u798f\u7530").strip(), "\u5176\u4ed6\u8f66\u578b\u540d\u5df2\u63d0\u53d6"
            elif why == "\u6cb9\u8f66":
                car, why, quality = "\u6cb9\u8f66\uff08\u8f66\u578b\u672a\u6ce8\u660e\uff09", "", "\u8f66\u578b\u4fe1\u606f\u4e0d\u8db3\uff1b\u539f\u56e0\u672a\u586b\u5199"
            else:
                car, quality = "\u5176\u4ed6\u8f66\u578b\uff08\u672a\u6ce8\u660e\uff09", "\u8f66\u578b\u4fe1\u606f\u4e0d\u8db3"
        out.append((model_name(car), car, why, quality))
    return out

def style(ws) -> None:
    ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
    fill = PatternFill("solid", fgColor="1E40AF")
    for c in ws[1]:
        c.fill, c.font, c.alignment = fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        values = [txt(ws.cell(r, col).value) for r in range(1, min(ws.max_row, 50) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(10, max(map(len, values), default=10) + 2))

def clean(source: Path, output: Path) -> None:
    src = load_workbook(source, read_only=True, data_only=True).active
    wb = Workbook(); main = wb.active; main.title = "\u6e05\u6d17\u6570\u636e"
    detail = wb.create_sheet("\u8f66\u578b\u539f\u56e0\u660e\u7ec6")
    main.append(CLEAN_HEADERS[:24] + ["\u5bf9\u6bd4\u8f66\u578b\uff08\u6807\u51c6\u6c47\u603b\uff09", "\u653e\u5f03\u539f\u56e0\uff08\u5bf9\u5e94\u987a\u5e8f\uff09", "\u8f66\u578b-\u539f\u56e0\u914d\u5bf9"] + CLEAN_HEADERS[24:])
    detail.append(["\u6837\u672c\u5e8f\u53f7", "\u624b\u673a\u53f7", "\u8f66\u578b\u5e8f\u53f7", "\u5bf9\u6bd4\u8f66\u578b\uff08\u6807\u51c6\uff09", "\u5bf9\u6bd4\u8f66\u578b\uff08\u539f\u59cb\uff09", "\u653e\u5f03\u539f\u56e0\uff08\u539f\u59cb\uff09", "\u653e\u5f03\u539f\u56e0\u7c7b\u522b", "\u89e3\u6790\u8d28\u91cf", "\u7b2c18\u9898\u539f\u59cb\u7b54\u6848"])
    rows = []
    for r in src.iter_rows(min_row=2, max_col=33, values_only=True):
        values = [clean_value(x) for x in r]
        pairs = parse(values[23]); missing = "\uff08\u672a\u586b\u5199\uff09"
        cars, whys = D.join(x[0] for x in pairs), D.join(x[2] or missing for x in pairs)
        paired = D.join(f"{x[0]}\u3016{x[2] or missing}\u3017" for x in pairs)
        main.append(values[:24] + [cars, whys, paired] + values[24:])
        for i, (car, raw_car, why, quality) in enumerate(pairs, 1):
            row = [values[0], values[6], i, car, raw_car, why, reason_type(why), quality, values[23]]
            rows.append(row); detail.append(row)
    summary = wb.create_sheet("\u8f66\u578bx\u539f\u56e0")
    cats = [x[0] for x in REASONS] + ["\u5176\u4ed6\u660e\u786e\u539f\u56e0", "\u672a\u8bf4\u660e/\u65e0\u6709\u6548\u539f\u56e0"]
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows: counts[row[3]][row[6]] += 1
    cars = sorted(counts, key=lambda x: (-sum(counts[x].values()), x))
    summary.append(["\u5bf9\u6bd4\u8f66\u578b\uff08\u6807\u51c6\uff09", *cats, "\u5408\u8ba1"])
    for car in cars:
        nums = [counts[car][cat] for cat in cats]; summary.append([car, *nums, sum(nums)])
    chart = BarChart(); chart.type = "bar"; chart.grouping = "stacked"; chart.overlap = 100
    chart.title = "\u4e3b\u8981\u5bf9\u6bd4\u8f66\u578b\u53ca\u653e\u5f03\u539f\u56e0\u6784\u6210"
    n = min(12, len(cars)); chart.add_data(Reference(summary, min_col=2, max_col=1+len(cats), min_row=1, max_row=n+1), titles_from_data=True)
    chart.set_categories(Reference(summary, min_col=1, min_row=2, max_row=n+1)); chart.height, chart.width = 11, 24
    summary.add_chart(chart, f"{get_column_letter(len(cats)+4)}2")
    note = wb.create_sheet("\u6e05\u6d17\u8bf4\u660e")
    notes = [
        ("\u9879\u76ee", "\u8bf4\u660e"), ("\u6837\u672c\u91cf", "153\u4efd\uff1b\u6e05\u6d17\u6570\u636e\u6bcf\u4f4d\u53d7\u8bbf\u8005\u4e00\u884c\u3002"),
        ("\u7b2c18\u9898\u62c6\u5206", "\u6309'\u250b'\u62c6\u8f66\u578b\uff0c\u6309\u7d27\u968f\u8f66\u578b\u540e\u7684'\u3016\u539f\u56e0\u3017'\u7ed1\u5b9a\u3002"),
        ("\u63a8\u8350\u5206\u6790\u8868", "\u8f66\u578b\u539f\u56e0\u660e\u7ec6\u4e3a\u4e00\u8f66\u4e00\u884c\uff0c\u662f\u4ea4\u53c9\u5206\u6790\u4e0e\u5236\u56fe\u7684\u6807\u51c6\u6570\u636e\u6e90\u3002"),
        ("\u7f3a\u5931\u5904\u7406", "\u4e0d\u63a8\u65ad\u672a\u586b\u539f\u56e0\uff0c\u7edf\u4e00\u5f52\u5165'\u672a\u8bf4\u660e/\u65e0\u6709\u6548\u539f\u56e0'\uff0c\u5e76\u6807\u8bb0\u89e3\u6790\u8d28\u91cf\u3002"),
        ("\u56fe\u8868", "\u8f66\u578bx\u539f\u56e0\u9875\u5c55\u793a\u524d12\u4e2a\u8f66\u578b\u7684\u5806\u79ef\u6761\u5f62\u56fe\uff1b\u6761\u957f\u770b\u5bf9\u6bd4\u9891\u6b21\uff0c\u989c\u8272\u770b\u539f\u56e0\u6784\u6210\u3002"),
    ]
    for row in notes: note.append(row)
    for ws in (main, detail, summary, note): style(ws)
    output.parent.mkdir(parents=True, exist_ok=True); wb.save(output)
    assert main.max_row == 154 and detail.max_row == len(rows)+1
    print(f"OK: {output} (153 samples, {len(rows)} model-reason rows)")

if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    clean(root/".data"/"\u4e94\u83f1\u626c\u5149Pro"/"381609494_\u6309\u6587\u672c_\u4e94\u83f1\u626c\u5149Pro\u7528\u6237\u8c03\u7814_153_153.xlsx", root/".data"/"\u4e94\u83f1\u626c\u5149Pro"/"\u4e94\u83f1\u626c\u5149Pro\u7528\u6237\u8c03\u7814153_\u6e05\u6d17\u7248.xlsx")
