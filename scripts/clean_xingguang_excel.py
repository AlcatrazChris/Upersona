"""Incrementally clean a Xingguang L survey using the verified 555-row result."""

from __future__ import annotations

import argparse
import re
import subprocess
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_REFERENCE = ROOT / ".data" / "wechat" / "2026-07" / "星光L首批用户调研_555.xlsx"
CLEAN_REFERENCE = ROOT / ".data" / "星光" / "星光L首批用户调研555.xlsx"
DELIMITER = "┋"
SKIP_VALUES = {"(跳过)", "（跳过）", "跳过"}
PROVINCE_REGION = {
    "黑龙江": "东北", "吉林": "东北", "辽宁": "东北", "内蒙古": "东北",
    "北京": "华北", "河北": "华北", "山西": "华北", "天津": "华北",
    "福建": "华东", "江苏": "华东", "上海": "华东", "浙江": "华东", "山东": "华东",
    "广东": "华南", "广西": "华南", "海南": "华南",
    "宁夏": "西北", "青海": "西北", "陕西": "西北", "新疆": "西北", "甘肃": "西北",
    "贵州": "西南", "四川": "西南", "云南": "西南", "重庆": "西南", "西藏": "西南",
    "安徽": "中南", "湖北": "中南", "湖南": "中南", "江西": "中南", "河南": "中南",
}

# cleaned column -> raw column
DIRECT_COLUMNS = {
    2: 6, 8: 9, 9: 10, 10: 11, 11: 12, 12: 18, 13: 19, 14: 21,
    15: 20, 16: 23, 17: 22, 18: 24, 19: 25, 20: 31, 21: 26,
    22: 27, 23: 28, 24: 29, 25: 30, 26: 46, 27: 55, 28: 39,
    29: 40, 30: 41, 31: 42, 33: 43, 34: 44, 37: 47, 38: 48,
    39: 53, 45: 63, 46: 64, 47: 65,
}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


# 2025 第一财经·新一线城市研究所《城市商业魅力排行榜》。项目将四、五线合并展示。
CITY_TIERS = {
    "一线城市": set("上海市 北京市 深圳市 广州市 新界".split()),
    "新一线城市": set("成都市 杭州市 重庆市 武汉市 苏州市 西安市 南京市 长沙市 郑州市 天津市 合肥市 青岛市 东莞市 宁波市 佛山市".split()),
    "二线城市": set("济南市 无锡市 沈阳市 昆明市 福州市 厦门市 温州市 石家庄市 大连市 哈尔滨市 金华市 泉州市 南宁市 长春市 常州市 南昌市 南通市 贵阳市 嘉兴市 徐州市 惠州市 太原市 烟台市 临沂市 保定市 台州市 绍兴市 珠海市 洛阳市 潍坊市".split()),
    "三线城市": set("乌鲁木齐市 兰州市 中山市 盐城市 海口市 扬州市 济宁市 湖州市 赣州市 邯郸市 南阳市 唐山市 芜湖市 阜阳市 廊坊市 汕头市 泰州市 呼和浩特市 镇江市 江门市 菏泽市 连云港市 沧州市 淄博市 新乡市 周口市 襄阳市 淮安市 商丘市 桂林市 咸阳市 上饶市 银川市 宿迁市 漳州市 遵义市 滁州市 绵阳市 宜昌市 威海市 湛江市 九江市 邢台市 揭阳市 三亚市 衡阳市 信阳市 泰安市 荆州市 肇庆市 蚌埠市 安阳市 安庆市 德州市 株洲市 莆田市 聊城市 驻马店市 岳阳市 亳州市 柳州市 宜春市 宿州市 黄冈市 六安市 常德市 宁德市 茂名市 马鞍山市 衢州市".split()),
}


def city_tier(city: Any) -> str:
    name = text(city)
    if not name:
        return ""
    return next((tier for tier, cities in CITY_TIERS.items() if name in cities), "四线城市及以下")


def learn_maps(raw_rows: list[tuple[Any, ...]], clean_rows: list[tuple[Any, ...]]):
    whole_counts = defaultdict(lambda: defaultdict(Counter))
    token_counts = defaultdict(lambda: defaultdict(Counter))
    for clean_col, raw_col in DIRECT_COLUMNS.items():
        for raw_row, clean_row in zip(raw_rows, clean_rows):
            raw, clean = text(raw_row[raw_col]), text(clean_row[clean_col])
            whole_counts[raw_col][raw][clean] += 1
            raw_parts, clean_parts = raw.split(DELIMITER), clean.split(DELIMITER)
            if len(raw_parts) == len(clean_parts):
                for source, target in zip(raw_parts, clean_parts):
                    token_counts[raw_col][source.strip()][target.strip()] += 1
    whole = {
        col: {raw: counts.most_common(1)[0][0] for raw, counts in values.items()}
        for col, values in whole_counts.items()
    }
    tokens = {
        col: {raw: counts.most_common(1)[0][0] for raw, counts in values.items()}
        for col, values in token_counts.items()
    }
    return whole, tokens


def fallback_token(value: str) -> str:
    value = value.strip()
    if not value or value in SKIP_VALUES:
        return ""
    value = re.sub(r"其他(?:原因)?\s*[〖【][^〗】]*[〗】]", "其他", value)
    return re.sub(r"\s*[〖【][^〗】]*[〗】]\s*", "", value).strip()


def clean_value(value: Any, raw_col: int, whole: dict, tokens: dict) -> str:
    raw = text(value)
    if not raw or raw in SKIP_VALUES:
        return ""
    if raw in whole.get(raw_col, {}):
        return whole[raw_col][raw]
    cleaned = []
    # Deliberately split on ┋ only. Slashes and punctuation belong to option text.
    for part in raw.split(DELIMITER):
        result = tokens.get(raw_col, {}).get(part.strip(), fallback_token(part))
        if result and result not in cleaned:
            cleaned.append(result)
    return DELIMITER.join(cleaned)


def learn_tuple_map(raw_rows, clean_rows, raw_col: int, clean_cols: tuple[int, ...]):
    counts = defaultdict(Counter)
    for raw_row, clean_row in zip(raw_rows, clean_rows):
        cleaned = tuple(text(clean_row[col]) for col in clean_cols)
        counts[text(raw_row[raw_col])][cleaned] += 1
    return {raw: values.most_common(1)[0][0] for raw, values in counts.items()}


def fallback_geo(value: Any, city_map: dict[str, tuple[str, ...]]) -> tuple[str, ...]:
    parts = [part.strip() for part in text(value).split("-")]
    province = parts[0] if parts else ""
    city = parts[1] if len(parts) > 1 else ""
    district = "-".join(parts[2:]) if len(parts) > 2 else ""
    known = city_map.get(city)
    return (known[0] if known else PROVINCE_REGION.get(province, ""), province, city, district, known[4] if known else "其他")


def fallback_budget(value: Any) -> tuple[str, str]:
    raw = text(value)
    numbers = re.findall(r"\d+(?:\.\d+)?", raw)
    if not numbers:
        return "", ""
    if "以上" in raw:
        return "", numbers[0]
    if len(numbers) >= 2:
        return numbers[1], numbers[0]
    return numbers[0], numbers[0]


def fallback_recommend(value: Any) -> tuple[str, str, str]:
    raw = text(value)
    detail_match = re.search(r"[〖【]([^〗】]*)[〗】]", raw)
    detail = detail_match.group(1).strip() if detail_match else ""
    if raw.startswith("推荐"):
        return "比较推荐", detail, ""
    if raw.startswith("不会推荐") or raw.startswith("不推荐"):
        return "不会推荐", "", detail
    if raw:
        return "说不准/看情况", "", ""
    return "", "", ""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    raw_ref_sheet = load_workbook(RAW_REFERENCE, read_only=True, data_only=True).active
    clean_ref_sheet = load_workbook(CLEAN_REFERENCE, read_only=True, data_only=True).active
    source_sheet = load_workbook(args.source, read_only=True, data_only=True).active
    raw_ref = list(raw_ref_sheet.iter_rows(min_row=2, values_only=True))
    clean_ref = list(clean_ref_sheet.iter_rows(min_row=2, values_only=True))
    source_rows = list(source_sheet.iter_rows(min_row=2, values_only=True))
    headers = list(next(clean_ref_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))

    if len(raw_ref) != len(clean_ref):
        raise ValueError("原始参考和清洗参考行数不一致")
    if any((text(a[1]), text(a[6])) != (text(b[0]), text(b[2])) for a, b in zip(raw_ref, clean_ref)):
        raise ValueError("参考数据无法按时间和姓名对齐")

    whole, tokens = learn_maps(raw_ref, clean_ref)
    geo_map = learn_tuple_map(raw_ref, clean_ref, 8, (3, 4, 5, 6, 7))
    city_map = {
        text(raw[8]).split("-")[1]: geo_map[text(raw[8])]
        for raw in raw_ref if len(text(raw[8]).split("-")) > 1
    }
    budget_map = learn_tuple_map(raw_ref, clean_ref, 45, (35, 36))
    vehicle_map = learn_tuple_map(raw_ref, clean_ref, 42, (32,))
    recommend_map = learn_tuple_map(raw_ref, clean_ref, 66, (48, 49, 50))

    output_rows = []
    for index, raw in enumerate(source_rows):
        if index < len(clean_ref) and (text(raw[1]), text(raw[6])) == (text(clean_ref[index][0]), text(clean_ref[index][2])):
            output_rows.append(list(clean_ref[index]))
            continue

        row = [""] * len(headers)
        row[0] = raw[1]
        for clean_col, raw_col in DIRECT_COLUMNS.items():
            row[clean_col] = clean_value(raw[raw_col], raw_col, whole, tokens)
        geo = list(geo_map.get(text(raw[8]), fallback_geo(raw[8], city_map)))
        row[3:8] = geo
        row[32] = vehicle_map.get(text(raw[42]), ("其他/无法识别",))[0]
        row[35], row[36] = budget_map.get(text(raw[45]), fallback_budget(raw[45]))
        row[48], row[49], row[50] = recommend_map.get(text(raw[66]), fallback_recommend(raw[66]))
        output_rows.append(row)

    for row in output_rows:
        row[7] = city_tier(row[5])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清洗数据"
    sheet.append(headers)
    for row in output_rows:
        sheet.append(row)
    workbook.save(args.output)

    subprocess.run(
        [
            "node", "--experimental-strip-types",
            str(ROOT / "scripts" / "reclassify-xingguang-vehicles.mjs"),
            str(args.source), str(args.output),
        ],
        cwd=ROOT,
        check=True,
    )
    final_column_count = load_workbook(args.output, read_only=True).active.max_column

    education_col = headers.index("学历")
    education = [text(row[education_col]) for row in output_rows if text(row[education_col])]
    if "高中/中专/职校/技校" not in education:
        raise AssertionError("学历复合选项被错误拆分")
    if any(value in SKIP_VALUES for row in output_rows for value in map(text, row)):
        raise AssertionError("清洗结果仍包含跳过值")
    valid_tiers = set(CITY_TIERS) | {"四线城市及以下"}
    if any(text(row[5]) and row[7] not in valid_tiers for row in output_rows):
        raise AssertionError("存在未分级城市")
    print(f"清洗完成：{args.output} ({len(output_rows)} 行 × {final_column_count} 列)")


if __name__ == "__main__":
    main()
