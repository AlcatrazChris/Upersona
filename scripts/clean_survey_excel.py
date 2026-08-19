"""Clean the raw Huajing S survey into the established analysis schema."""

from __future__ import annotations

import argparse
import re
import subprocess
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from clean_xingguang_excel import city_tier


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / ".data" / "wechat" / "2026-07" / "366639047_3644_3644.xlsx"
REFERENCE = ROOT / ".data" / "git-history" / "华境S首批用户调研0617.xlsx"
OUTPUT = ROOT / ".data" / "华境" / "华境S首批用户调研3644.xlsx"

DELIMITER = "┋"
SKIP_VALUES = {"(跳过)", "（跳过）", "跳过"}

# (raw source column, cleaned reference column), zero based.
DIRECT_COLUMNS = (
    [(6, 1)]
    + [(source_col, source_col - 2) for source_col in range(9, 33)]
    + [(source_col, source_col - 1) for source_col in range(33, 51)]
)

OTHER_RE = re.compile(r"其他(?:原因)?\s*[〖【][^〗】]*[〗】]")
EXPLANATION_RE = re.compile(r"\s*[（(][^（）()]*[）)]\s*")


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_time(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    raw = text(value)
    if not raw:
        return None
    for fmt in (
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def build_clean_maps(
    source_rows: list[tuple[Any, ...]],
    reference_rows: list[tuple[Any, ...]],
) -> tuple[dict[int, dict[str, str]], dict[int, dict[str, str]]]:
    """Learn the exact cleaning vocabulary from the aligned reference rows."""
    whole_counts: dict[int, dict[str, Counter[str]]] = defaultdict(
        lambda: defaultdict(Counter)
    )
    token_counts: dict[int, dict[str, Counter[str]]] = defaultdict(
        lambda: defaultdict(Counter)
    )

    aligned_count = min(len(source_rows), len(reference_rows))
    for source_col, target_col in DIRECT_COLUMNS:
        for index in range(aligned_count):
            raw = text(source_rows[index][source_col])
            clean = text(reference_rows[index][target_col])
            whole_counts[source_col][raw][clean] += 1

            raw_parts = raw.split(DELIMITER)
            clean_parts = clean.split(DELIMITER)
            if len(raw_parts) == len(clean_parts):
                for raw_part, clean_part in zip(raw_parts, clean_parts):
                    token_counts[source_col][raw_part.strip()][clean_part.strip()] += 1

    whole_maps = {
        column: {raw: counts.most_common(1)[0][0] for raw, counts in mappings.items()}
        for column, mappings in whole_counts.items()
    }
    token_maps = {
        column: {raw: counts.most_common(1)[0][0] for raw, counts in mappings.items()}
        for column, mappings in token_counts.items()
    }
    return whole_maps, token_maps


def fallback_clean_token(value: str) -> str:
    value = value.strip()
    if not value or value in SKIP_VALUES:
        return ""
    value = OTHER_RE.sub("其他", value)
    # Explanatory parentheses in questionnaire options are descriptions,
    # not part of the category name.
    value = EXPLANATION_RE.sub("", value).strip()
    # Work/life-state options use "category: long description".
    if ":" in value or "：" in value:
        value = re.split(r"[:：]", value, maxsplit=1)[0].strip()
    return value


def normalize_industry(*values: Any) -> str:
    primary = text(values[0]) if values else ""
    canonical = {
        "公共管理与事业单位", "制造与工业服务", "商贸零售与个体经营",
        "建筑与房地产", "教育培训", "信息技术与互联网通信",
        "住宿餐饮与生活服务", "医疗健康与社会服务", "金融与财会",
        "能源矿产与公用事业", "交通物流与邮政", "科研与专业服务",
        "农林牧渔", "文化体育娱乐与旅游", "退休", "租赁服务", "其他",
    }
    if primary in canonical - {"其他"}:
        return primary
    rules = (
        (r"退休|离休|退修|已退", "退休"),
        (r"政府|事业单位|国企|国有企业|国有/集体企业|检察院|军队|部队|社区|社会组织|香港政府|退役军人|国际组织", "公共管理与事业单位"),
        (r"建筑|房地产|装修|装饰|物业", "建筑与房地产"),
        (r"批发|零售|电商|外贸|贸易|销售|商业|酒类|彩票|手机批发|小买卖|商人|做生意|经营者|个体|创业|自营|合伙人|市场管理|区域经理|开店|投资|促销", "商贸零售与个体经营"),
        (r"教育|培训|教师", "教育培训"),
        (r"医疗|卫生|社会保障|社会福利|医生|护士|药师|检验师|药品|生物实验|医疗器械", "医疗健康与社会服务"),
        (r"金融|财务|会计", "金融与财会"),
        (r"计算机|软件|互联网|通信|通讯|电信|物联网|自媒体|AI|光通信", "信息技术与互联网通信"),
        (r"物流|交通|仓储|邮政|司机|客运|公路|高速|收费员|外卖|校车", "交通物流与邮政"),
        (r"住宿|餐饮|生活服务|美容|美发|洗护|家政|服务业|服务〗", "住宿餐饮与生活服务"),
        (r"农、林、牧、渔|农民", "农林牧渔"),
        (r"电力|燃气|水生产|水利|环境|公共设施|采矿|能源|石油|化工|炼油|石化|危化|新能源|柴汽油", "能源矿产与公用事业"),
        (r"科学研究|技术服务|地质勘查|专业服务|法律|咨询|设计服务|广告服务|中介", "科研与专业服务"),
        (r"文化|体育|娱乐|旅游|媒体|作家|记者|设计师|演员|摄影|舞美", "文化体育娱乐与旅游"),
        (r"出租|租赁", "租赁服务"),
        (r"制造|生产|加工|工程师|技术员|研发人员|操作人员|操作工|汽车|汽配|纺织|服装|家具|家电|食品工厂|维修|汽修|机动车检测|弱电工|电工|技术工|工人|飞机维修|建材|五金|试驾车", "制造与工业服务"),
    )
    for pattern, category in rules:
        if re.search(pattern, primary):
            return category
    useful = [
        value for value in map(text, values)
        if value and value not in SKIP_VALUES
        and not re.fullmatch(r"(?:其他[〖【]?)?(?:无|无业|待业|没有工作|不便透露|秘密|其他)[〗】]?", value)
    ]
    if not useful or all(re.search(r"没有工作|主妇|主夫|学生|无业|待业", value) for value in useful):
        return ""
    support = [text(value) for value in values[1:] if text(value) not in SKIP_VALUES]
    support_raw = " ".join(support)
    for pattern, category in rules:
        haystack = support[0] if category == "公共管理与事业单位" and support else support_raw
        if re.search(pattern, haystack):
            return category
    return "其他"


def clean_cell(
    value: Any,
    source_col: int,
    whole_maps: dict[int, dict[str, str]],
    token_maps: dict[int, dict[str, str]],
) -> str:
    raw = text(value)
    if not raw or raw in SKIP_VALUES:
        return ""
    if raw in whole_maps.get(source_col, {}):
        return whole_maps[source_col][raw]

    cleaned: list[str] = []
    for part in raw.split(DELIMITER):
        part = part.strip()
        result = token_maps.get(source_col, {}).get(part)
        if result is None:
            result = fallback_clean_token(part)
        if result and result not in cleaned:
            cleaned.append(result)
    return DELIMITER.join(cleaned)


def build_geo_maps(
    source_rows: list[tuple[Any, ...]],
    reference_rows: list[tuple[Any, ...]],
) -> tuple[dict[str, tuple[str, ...]], dict[str, tuple[str, ...]], dict[str, str]]:
    exact: dict[str, tuple[str, ...]] = {}
    city: dict[str, tuple[str, ...]] = {}
    province_region: dict[str, str] = {}

    for source_row, reference_row in zip(source_rows, reference_rows):
        raw = text(source_row[8])
        geo = tuple(text(value) for value in reference_row[2:7])
        if not raw:
            continue
        exact[raw] = geo
        parts = [part.strip() for part in raw.split("-")]
        if len(parts) >= 2:
            city[parts[1]] = geo
        if geo[1] and geo[0]:
            province_region[geo[1]] = geo[0]
    return exact, city, province_region


def resolve_geo(
    value: Any,
    exact_map: dict[str, tuple[str, ...]],
    city_map: dict[str, tuple[str, ...]],
    province_region: dict[str, str],
) -> tuple[str, str, str, str, str]:
    raw = text(value)
    if raw in exact_map:
        geo = exact_map[raw]
        return (*geo[:4], geo[4] or city_tier(geo[2]))  # type: ignore[return-value]

    parts = [part.strip() for part in raw.split("-")]
    province = parts[0] if parts else ""
    city = parts[1] if len(parts) > 1 else ""
    district = "-".join(parts[2:]) if len(parts) > 2 else ""
    known = city_map.get(city)
    if known:
        resolved_city = city or known[2]
        return known[0], province or known[1], resolved_city, district, known[4] or city_tier(resolved_city)
    return province_region.get(province, ""), province, city, district, city_tier(city)


def build_car_map(
    source_rows: list[tuple[Any, ...]],
    reference_rows: list[tuple[Any, ...]],
) -> dict[str, str]:
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    for source_row, reference_row in zip(source_rows, reference_rows):
        raw = text(source_row[32]).lower()
        category = text(reference_row[31])
        counts[raw][category] += 1
    return {raw: values.most_common(1)[0][0] for raw, values in counts.items()}


def classify_car(value: Any, known_map: dict[str, str]) -> str:
    raw = text(value)
    if not raw or raw in SKIP_VALUES:
        return ""
    lowered = raw.lower()
    if lowered in known_map:
        return known_map[lowered]

    if any(keyword in lowered for keyword in ("五菱", "宝骏")):
        return "五菱宝骏基盘"
    if any(keyword in lowered for keyword in (
        "问界", "理想", "蔚来", "小鹏", "零跑", "哪吒", "极氪", "阿维塔", "岚图",
    )):
        return "新势力"
    if any(keyword in lowered for keyword in (
        "比亚迪", "埃安", "深蓝", "银河", "欧拉", "启源", "腾势",
    )):
        return "国产新能源"
    if any(keyword in lowered for keyword in (
        "奔驰", "宝马", "奥迪", "保时捷", "路虎", "捷豹", "雷克萨斯",
        "凯迪拉克", "沃尔沃", "林肯", "英菲尼迪", "讴歌", "特斯拉",
    )):
        return "合资豪华"
    if any(keyword in lowered for keyword in (
        "吉利", "长安", "奇瑞", "传祺", "红旗", "哈弗", "领克", "名爵",
        "荣威", "江淮", "东风", "奔腾", "北京汽车", "北汽", "广汽",
    )):
        return "国产燃油"
    return "合资主流"


def copy_header_style(reference_sheet: Any, output_sheet: Any, offset: int = 1) -> None:
    for source_cell in reference_sheet[1]:
        target_cell = output_sheet.cell(1, source_cell.column + offset)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def main() -> None:
    assert city_tier("北京市") == "一线城市"
    assert city_tier("未收录市") == "四线城市及以下"
    parser = argparse.ArgumentParser()
    parser.add_argument("source", nargs="?", type=Path, default=SOURCE)
    parser.add_argument("output", nargs="?", type=Path, default=OUTPUT)
    args = parser.parse_args()

    source_book = load_workbook(args.source, read_only=True, data_only=True)
    reference_book = load_workbook(REFERENCE, data_only=True)
    source_sheet = source_book.active
    reference_sheet = reference_book.active

    source_rows = list(source_sheet.iter_rows(min_row=2, values_only=True))
    reference_rows = list(reference_sheet.iter_rows(min_row=2, values_only=True))

    whole_maps, token_maps = build_clean_maps(source_rows, reference_rows)
    exact_geo, city_geo, province_region = build_geo_maps(source_rows, reference_rows)
    car_map = build_car_map(source_rows, reference_rows)

    output_book = Workbook()
    output_sheet = output_book.active
    output_sheet.title = "清洗数据"

    reference_headers = [cell.value for cell in reference_sheet[1]]
    output_sheet.append(["提交答卷时间", *reference_headers])
    copy_header_style(reference_sheet, output_sheet)
    # Match the target header style for the added time field.
    first_reference_header = reference_sheet.cell(1, 1)
    output_sheet.cell(1, 1)._style = copy(first_reference_header._style)
    output_sheet.cell(1, 1).font = copy(first_reference_header.font)
    output_sheet.cell(1, 1).fill = copy(first_reference_header.fill)
    output_sheet.cell(1, 1).border = copy(first_reference_header.border)
    output_sheet.cell(1, 1).alignment = copy(first_reference_header.alignment)

    output_count = 0
    removed_blank_rows = 0
    removed_skip_cells = 0
    other_normalized = 0

    for source_row in source_rows:
        removed_skip_cells += sum(text(value) in SKIP_VALUES for value in source_row)
        other_normalized += sum(
            bool(OTHER_RE.search(text(value))) for value in source_row if value is not None
        )

        cleaned_direct: dict[int, str] = {}
        for source_col, _ in DIRECT_COLUMNS:
            cleaned_direct[source_col] = clean_cell(
                source_row[source_col], source_col, whole_maps, token_maps
            )
        cleaned_direct[13] = normalize_industry(*(source_row[column] for column in range(13, 17)))

        # A row is blank only when all actual survey answers are empty after cleaning.
        if not any(cleaned_direct.values()):
            removed_blank_rows += 1
            continue

        geo = resolve_geo(source_row[8], exact_geo, city_geo, province_region)
        row = [
            parse_time(source_row[1]),
            "已提车",
            cleaned_direct[6],
            *geo,
            *[cleaned_direct[column] for column in range(9, 33)],
            classify_car(source_row[32], car_map),
            *[cleaned_direct[column] for column in range(33, 51)],
        ]
        output_sheet.append(row)
        output_count += 1
        output_sheet.cell(output_sheet.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"

    output_sheet.freeze_panes = "A2"
    output_sheet.auto_filter.ref = output_sheet.dimensions
    output_sheet.column_dimensions["A"].width = 21
    for column_index in range(1, reference_sheet.max_column + 1):
        source_letter = reference_sheet.cell(1, column_index).column_letter
        target_letter = output_sheet.cell(1, column_index + 1).column_letter
        output_sheet.column_dimensions[target_letter].width = (
            reference_sheet.column_dimensions[source_letter].width or 13
        )
    output_sheet.row_dimensions[1].height = reference_sheet.row_dimensions[1].height

    args.output.parent.mkdir(parents=True, exist_ok=True)
    output_book.save(args.output)
    subprocess.run(
        [
            "node", "--experimental-strip-types",
            str(ROOT / "scripts" / "reclassify-xingguang-vehicles.mjs"),
            str(args.source), str(args.output),
        ],
        cwd=ROOT,
        check=True,
    )
    print(f"输出文件: {args.output}")
    print(f"源数据行: {len(source_rows)}")
    print(f"输出数据行: {output_count}")
    print(f"删除全空行: {removed_blank_rows}")
    print(f"清除跳过单元格: {removed_skip_cells}")
    print(f"归并其他选项: {other_normalized}")


if __name__ == "__main__":
    main()
